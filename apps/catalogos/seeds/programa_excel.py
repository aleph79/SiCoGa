"""Reads the PROGRAMA REIMPLANTES sheet and yields normalized rows.

Excel layout: 5 sections with header rows: MACHOS, HEMBRAS, GANADO DE POTRERO MACHO,
GANADO DE POTRERO HEMBRA, VACAS. Each section has its own column structure.

This loader returns a flat list of dicts with the canonical fields used by the
ProgramaReimplante model.
"""

from pathlib import Path

from openpyxl import load_workbook

SECCIONES = {
    "MACHOS": {"tipo_ganado": "Macho", "tipo_origen": "Corral"},
    "HEMBRAS": {"tipo_ganado": "Hembra", "tipo_origen": "Corral"},
    "GANADO DE POTRERO MACHO": {"tipo_ganado": "Macho", "tipo_origen": "Potrero"},
    "GANADO DE POTRERO HEMBRA": {"tipo_ganado": "Hembra", "tipo_origen": "Potrero"},
    "VACAS": {"tipo_ganado": "Vaca", "tipo_origen": None},
}


def _parse_rango(s):
    """'100-150' -> (100, 150). 'MENOR A 400' -> (0, 400). 'MAYOR A 400' -> (400.01, 9999)."""
    if s is None:
        return None, None
    s = str(s).strip()
    upper = s.upper()
    if "MENOR A" in upper:
        return 0, int(upper.split("MENOR A")[1].strip())
    if "MAYOR A" in upper:
        base = int(upper.split("MAYOR A")[1].strip())
        return base + 0.01, 9999
    if s.endswith("+"):
        return int(s[:-1]), 9999
    if "-" in s:
        a, b = s.split("-")
        return int(a), int(b)
    return None, None


def _clean_implante(value):
    if not isinstance(value, str):
        return ""
    s = value.strip()
    if not s or s.upper() == "N/A":
        return ""
    return s[:40]


def leer_programa_excel(path: Path):
    wb = load_workbook(str(path), data_only=True)
    ws = wb["PROGRAMA REIMPLANTES"]
    filas = []
    seccion_actual = None

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        first = row[0]
        if isinstance(first, str) and first.strip().upper() in SECCIONES:
            seccion_actual = first.strip().upper()
            continue
        if seccion_actual is None or first is None:
            continue
        if isinstance(first, str) and first.upper().startswith(("RANGO", "PROYECCION", "PROM")):
            continue
        peso_min, peso_max = _parse_rango(first)
        if peso_min is None:
            continue

        meta = SECCIONES[seccion_actual]

        # Layout machos/hembras/potrero: [rango, prom, kg_hacer, gdp, dias_estancia,
        #   implante, reimp1, reimp2, reimp3, reimp4, recepcion, dias_f1, dias_trans,
        #   dias_f3, dias_z, total_dias, peso_salida]
        gdp = row[3] if len(row) > 3 else None
        implante = row[5] if len(row) > 5 else None
        reimp1 = row[6] if len(row) > 6 else None
        reimp2 = row[7] if len(row) > 7 else None
        reimp3 = row[8] if len(row) > 8 else None
        reimp4 = row[9] if seccion_actual == "MACHOS" and len(row) > 9 else None
        recepcion = row[10] if len(row) > 10 else 0
        dias_f1 = row[11] if len(row) > 11 else 0
        dias_trans = row[12] if len(row) > 12 else 14
        dias_f3 = row[13] if len(row) > 13 else 0
        dias_z = row[14] if len(row) > 14 else 35
        peso_salida = row[16] if len(row) > 16 else 580

        if gdp is None:
            continue

        filas.append(
            {
                "categoria": seccion_actual,
                "tipo_ganado": meta["tipo_ganado"],
                "tipo_origen": meta["tipo_origen"],
                "peso_min": peso_min,
                "peso_max": peso_max,
                "gdp_esperada": gdp,
                "peso_objetivo_salida": peso_salida or 580,
                "implante_inicial": _clean_implante(implante),
                "reimplante_1": _clean_implante(reimp1),
                "reimplante_2": _clean_implante(reimp2),
                "reimplante_3": _clean_implante(reimp3),
                "reimplante_4": _clean_implante(reimp4),
                "dias_recepcion": int(recepcion or 0),
                "dias_f1": int(dias_f1 or 0),
                "dias_transicion": int(dias_trans or 14),
                "dias_f3": int(dias_f3 or 0),
                "dias_zilpaterol": int(dias_z or 35),
            }
        )
    return filas
